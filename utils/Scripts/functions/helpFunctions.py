import re
import unicodedata
from difflib import SequenceMatcher
import openpyxl
from tabulate import tabulate
import os
import tkinter as tk
from tkinter import filedialog

# ==============================================================================
#    FUNCIONES DE AYUDA (LOGICA DE NORMALIZACION Y COMPARACION)
# ==============================================================================

def numero_a_romano(match):
    """Callback para convertir números arábigos (1-10) a romanos."""
    mapa = {
        '1': 'I', '2': 'II', '3': 'III', '4': 'IV', '5': 'V',
        '6': 'VI', '7': 'VII', '8': 'VIII', '9': 'IX', '10': 'X'
    }
    return mapa.get(match.group(), match.group())

def normalizar_para_comparacion(texto):
    if not texto: return ""
    
    # 1. Limpieza básica
    texto = str(texto).lower().strip()

    # --- CORRECCIÓN: ELIMINAMOS EL BORRADO AGRESIVO DE PARÉNTESIS ---
    # Ya no usamos re.sub(r'\(.*?\)', '', texto) porque borraba partes útiles del nombre.
    # La limpieza de (*) ya se hizo previamente con limpiar_nombre_asignatura
    # ---------------------------------------------------------------
    
    # 2. Convertir número final a romano
    # Nota: Validamos que el texto termine en un número aislado
    texto = re.sub(r'\b([1-9]|10)\b$', numero_a_romano, texto, flags=re.IGNORECASE)
    
    # 3. Quitar acentos (NFD)
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    
    # 4. Eliminar espacios para crear una "huella digital" del texto
    return texto.replace(" ", "")

def extraer_numeros_clave(texto):
    """Extrae TODOS los números (arábigos y romanos) para distinguir niveles."""
    texto = str(texto).upper()
    arabigos = re.findall(r'\b\d+\b', texto)
    romanos = re.findall(r'\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b', texto)
    return set(arabigos + romanos)

def es_parecido(nombre_nuevo, nombre_existente, umbral=0.92):
    """Comparación visual difusa (Levenshtein ratio)."""
    return SequenceMatcher(None, nombre_nuevo, nombre_existente).ratio() >= umbral

def procesar_excel_exacto(archivo, nombre_hoja, indices_columnas, fila_inicio):
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        if nombre_hoja not in wb.sheetnames:
            print(f"Error: No existe la hoja '{nombre_hoja}'")
            return
        
        sheet = wb[nombre_hoja]
        data_final = []

        for num_fila, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            # Saltamos hasta llegar a la fila deseada
            if num_fila < fila_inicio:
                continue

            # CORRECCIÓN 1: Inicializamos vacío, sin el num_fila
            fila_procesada = [] 

            # Extraer las columnas específicas por índice
            for i in indices_columnas:
                if i < len(row):
                    fila_procesada.append(row[i])
                else:
                    fila_procesada.append(None)

            # CORRECCIÓN 2: Validamos toda la fila (ya no usamos [1:])
            # Solo agregamos si hay contenido real
            if any(celda is not None for celda in fila_procesada):
                data_final.append(fila_procesada)

        # Mostrar con formato
        if data_final:
            # CORRECCIÓN 3: Quitamos el encabezado de "# Fila"
            '''headers = [f"Col {i}" for i in indices_columnas]
            print(f"\nDatos de '{nombre_hoja}' (Desde fila {fila_inicio}):")
            print(tabulate(data_final, headers=headers, tablefmt="fancy_grid"))'''
            
            return data_final
        else:
            print(f"No hay datos a partir de la fila {fila_inicio}")

    except Exception as e:
        print(f"Error crítico: {e}")

def seleccion_archivo():
    root = tk.Tk()
    root.withdraw() # Ocultamos la ventanita principal

    # TRUCO: Forzar que la ventana aparezca al frente de todo
    root.attributes('-topmost', True) 
    
    print("📂 Abriendo ventana de selección...")
    
    ruta_archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de Malla",
        filetypes=[
            ("Archivos de Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    
    # Importante: Destruimos la instancia de tkinter para liberar memoria
    root.destroy() 

    # Validamos
    if ruta_archivo:
        nombre_archivo = os.path.basename(ruta_archivo)
        print(f"✅ Archivo seleccionado: {nombre_archivo}")
        return ruta_archivo
    else:
        print("❌ No se seleccionó ningún archivo.")
        return None

def obtener_nombre_hoja(ruta_archivo):
    """Muestra las hojas disponibles y permite elegir una."""
    try:
        # Leemos solo la estructura para ser rápidos
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, keep_links=False)
        hojas = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"❌ Error al leer hojas: {e}")
        return None

    if len(hojas) == 1:
        print(f"✅ Hoja única detectada: {hojas[0]}")
        return hojas[0]

    print("\n📋 Hojas disponibles:")
    for i, hoja in enumerate(hojas):
        print(f"  [{i + 1}] {hoja}")

    while True:
        try:
            opcion = int(input("\n👉 Elige el número de la hoja: ")) - 1
            if 0 <= opcion < len(hojas):
                return hojas[opcion]
            print("⚠️ Número inválido.")
        except ValueError:
            print("⚠️ Ingresa un número.")

def obtener_año_periodo(archivo, nombre_hoja):
    direccion = 'D1'
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        

        ws = wb[nombre_hoja]
        año_periodo = ws[direccion].value
        wb.close()

        # Si la celda está vacía, devolvemos valores por defecto
        if not año_periodo:
            return False, False

        texto_limpio = str(año_periodo).upper()
        
        # 1. AÑO (Protegido)
        match = re.search(r'\d{4}', texto_limpio)
        if match:
            year = int(match.group())
        else:
            return False, False
            
        # 2. PERIODO (Con valor por defecto)
        periodo = 1 # Valor predeterminado (asumimos 1 si no dice nada)
        
        if "SEGUNDO" in texto_limpio or "2DO" in texto_limpio:
            periodo = 2
        elif "PRIMER" in texto_limpio or "1ER" in texto_limpio:
            periodo = 1
        else:
            return False, False
            
        return year, periodo

    except Exception as e:
        print(f"❌ Error procesando fechas: {e}")
        return False, False

def limpiar_pantalla():
    # Funciona en Windows (cls) y Linux/Mac (clear)
    os.system('cls' if os.name == 'nt' else 'clear')
    
def formatDoc(nom, app):
    """
    Formatea el nombre y apellido para obtener el primer nombre y primer apellido.
    """
    # Convertir a string y quitar espacios extra
    nombre_raw = str(nom).strip() if nom else ""
    apellido_raw = str(app).strip() if app else ""
    
    # Validación básica
    if not nombre_raw or not apellido_raw:
        print(f"⚠️ Docente omitido por datos incompletos: {nom}, {app}")
        return False
    
    # Extraer el primer nombre y el primer apellido
    try:
        nom_pila = nombre_raw.split()[0]
        ape_pila = apellido_raw.split()[0]
        
        # Retornar "Nombre Apellido"
        return f"{nom_pila} {ape_pila}"
    except IndexError:
        return False

def limpiar_nombre_asignatura(texto):
    """
    Recibe un nombre sucio (ej: 'Matemática I (*)') y devuelve el nombre limpio
    listo para ser procesado o guardado (ej: 'Matemática I').
    """
    if not texto: 
        return ""
    
    texto = str(texto)
    
    # 1. Eliminar marcadores al final como (*), (**), (***) usando Regex
    texto = re.sub(r'\s*\(\*+\)$', '', texto) 
    
    # 2. Eliminación literal por seguridad (si aparecen en medio del texto)
    texto = texto.replace("(*)", "").replace("(**)", "").replace("(***)", "")
    
    # 3. Quitar espacios múltiples y extremos
    return " ".join(texto.split())