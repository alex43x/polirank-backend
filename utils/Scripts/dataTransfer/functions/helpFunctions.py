import os
import openpyxl
import re
from tabulate import tabulate
import tkinter as tk
from tkinter import filedialog
import unicodedata
from typing import List, Optional, Any


def limpiar_pantalla():
    os.system('cls' if os.name == 'nt' else 'clear')
    
    
def procesar_excel_exacto(archivo, nombre_hoja, indices_columnas, fila_inicio):
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        if nombre_hoja not in wb.sheetnames:
            print(f"Error: No existe la hoja '{nombre_hoja}'")
            return
        
        sheet = wb[nombre_hoja]
        data_final = []

        for num_fila, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            # Saltamos hasta llegar a la fila deseada
            if num_fila < fila_inicio:
                continue

            # CORRECCIÓN 1: Inicializamos vacío, sin el num_fila
            fila_procesada = [] 

            # Extraer las columnas específicas por índice
            for i in indices_columnas:
                if i < len(row):
                    fila_procesada.append(row[i])
                else:
                    fila_procesada.append(None)

            # CORRECCIÓN 2: Validamos toda la fila (ya no usamos [1:])
            # Solo agregamos si hay contenido real
            if any(celda is not None for celda in fila_procesada):
                data_final.append(fila_procesada)

        # Mostrar con formato
        if data_final:
            # CORRECCIÓN 3: Quitamos el encabezado de "# Fila"
            '''headers = [f"Col {i}" for i in indices_columnas]
            print(f"\nDatos de '{nombre_hoja}' (Desde fila {fila_inicio}):")
            print(tabulate(data_final, headers=headers, tablefmt="fancy_grid"))'''
            return data_final
        else:
            print(f"No hay datos a partir de la fila {fila_inicio}")
            return []

    except Exception as e:
        print(f"Error crítico: {e}")
        
        
def seleccion_archivo():
    root = tk.Tk()
    root.withdraw() # Ocultamos la ventanita principal

    # TRUCO: Forzar que la ventana aparezca al frente de todo
    root.attributes('-topmost', True) 
    
    print("📂 Abriendo ventana de selección...")
    
    ruta_archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de Malla",
        filetypes=[
            ("Archivos de Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    
    # Importante: Destruimos la instancia de tkinter para liberar memoria
    root.destroy() 

    # Validamos
    if ruta_archivo:
        nombre_archivo = os.path.basename(ruta_archivo)
        print(f"✅ Archivo seleccionado: {nombre_archivo}")
        return ruta_archivo
    else:
        print("❌ No se seleccionó ningún archivo.")
        return None
    
    
def obtener_nombre_hoja(ruta_archivo):
    """Muestra las hojas disponibles y permite elegir una."""
    try:
        # Leemos solo la estructura para ser rápidos
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, keep_links=False)
        hojas = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"❌ Error al leer hojas: {e}")
        return None

    if len(hojas) == 1:
        print(f"✅ Hoja única detectada: {hojas[0]}")
        return hojas[0]

    print("\n📋 Hojas disponibles:")
    for i, hoja in enumerate(hojas):
        print(f"  [{i + 1}] {hoja}")

    while True:
        try:
            opcion = int(input("\n👉 Elige el número de la hoja: ")) - 1
            if 0 <= opcion < len(hojas):
                return hojas[opcion]
            print("⚠️ Número inválido.")
        except ValueError:
            print("⚠️ Ingresa un número.")
            
            
def formatDoc(nom, app):
    """Une Nombre y Apellido completos."""
    nombre_raw = str(nom).strip() if nom else ""
    apellido_raw = str(app).strip() if app else ""
    
    if not nombre_raw or not apellido_raw:
        return False
    
    try:
        # ANTES: Solo primer nombre + primer apellido
        # nom_pila = nombre_raw.split()[0]
        # ape_pila = apellido_raw.split()[0]
        # return f"{nom_pila} {ape_pila}"
        
        # AHORA: Nombre completo + Apellido completo
        nombre_completo = " ".join(nombre_raw.split())  # Normaliza espacios
        apellido_completo = " ".join(apellido_raw.split())  # Normaliza espacios
        
        return f"{nombre_completo} {apellido_completo}"
        
    except Exception as e:
        print(f"❌ Error formateando nombre completo: {e}")
        return False
    
def normalizar_nombre_comparacion(nombre):
    """
    Normaliza nombre para comparación (quita tildes, convierte a minúsculas, etc.)
    """
    if not nombre:
        return ""
    
    # Convertir a minúsculas
    nombre = nombre.lower()
    
    # Quitar tildes
    nombre = unicodedata.normalize('NFD', nombre)
    nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')
    
    # Quitar espacios extra
    nombre = " ".join(nombre.split())
    
    return nombre

def nombres_similares(nombre1, nombre2):
    """
    Compara si dos nombres son similares.
    CORRECCIÓN: Ahora es más estricto - compara el nombre completo normalizado.
    Solo considera duplicado si el nombre completo normalizado es igual o muy similar.
    """
    if not nombre1 or not nombre2:
        return False
    
    # Normalizar ambos nombres completamente
    n1 = normalizar_nombre_comparacion(nombre1)
    n2 = normalizar_nombre_comparacion(nombre2)
    
    # Si son exactamente iguales después de normalizar, son la misma persona
    if n1 == n2:
        return True
    
    # CORRECCIÓN: En lugar de solo comparar primer nombre + primer apellido,
    # comparamos si tienen al menos 3 palabras en común (para evitar falsos positivos)
    palabras1 = set(n1.split())
    palabras2 = set(n2.split())
    
    # Si tienen menos de 2 palabras, no son similares
    if len(palabras1) < 2 or len(palabras2) < 2:
        return False
    
    # Calcular intersección de palabras
    palabras_comunes = palabras1.intersection(palabras2)
    
    # CORRECCIÓN: Requerir al menos 2 palabras en común Y que coincidan primer nombre y primer apellido
    if len(palabras_comunes) >= 2:
        palabras1_list = n1.split()
        palabras2_list = n2.split()
        # Verificar que coincidan primer nombre Y primer apellido
        if (len(palabras1_list) >= 2 and len(palabras2_list) >= 2 and
            palabras1_list[0] == palabras2_list[0] and 
            palabras1_list[-1] == palabras2_list[-1]):
            
            # CORRECCIÓN 2: Si ambos tienen 3+ palabras y difieren en alguna del medio, son personas distintas
            # Ejemplo: "José Antonio González" vs "José María González" -> NO son la misma persona
            if len(palabras1_list) >= 3 and len(palabras2_list) >= 3:
                # Si tienen la misma cantidad de palabras, verificar que todas las del medio coincidan
                if len(palabras1_list) == len(palabras2_list):
                    # Comparar palabras del medio (excluyendo primera y última)
                    palabras_medio_1 = palabras1_list[1:-1]
                    palabras_medio_2 = palabras2_list[1:-1]
                    # Si hay diferencias en las palabras del medio, son personas distintas
                    if palabras_medio_1 != palabras_medio_2:
                        return False
            
            return True
    
    return False

def extraer_nombre_clave(nombre_completo):
    """
    Extrae el nombre clave para comparación: primer nombre + primer apellido
    """
    if not nombre_completo:
        return ""
    
    nombre_normalizado = normalizar_nombre_comparacion(nombre_completo)
    palabras = nombre_normalizado.split()
    
    if len(palabras) >= 2:
        primer_nombre = palabras[0]
        primer_apellido = palabras[-1]
        return f"{primer_nombre} {primer_apellido}"
    
    return nombre_normalizado

def detectar_duplicados_por_nombre(nuevos_docentes):
    """
    Detecta duplicados por nombre.
    """
    nombres_vistos = {}
    duplicados = []
    
    for i, (nombre, correo) in enumerate(nuevos_docentes):
        encontrado = False
        nombre_clave = extraer_nombre_clave(nombre)
        
        # Buscar por nombre clave
        if nombre_clave in nombres_vistos:
            datos = nombres_vistos[nombre_clave]
            if nombres_similares(nombre, datos['nombre']):
                duplicados.append({
                    'indice_actual': i,
                    'nombre': nombre,
                    'correo_actual': correo,
                    'duplicado_de': datos['indice'],
                    'nombre_duplicado': datos['nombre'],
                    'correo_duplicado': datos['correo']
                })
                encontrado = True
        
        if not encontrado:
            nombres_vistos[nombre_clave] = {
                'indice': i,
                'nombre': nombre,
                'correo': correo
            }
    
    return duplicados

def preferir_correo_institucional(correo1, correo2):
    """
    Decide cuál correo preferir: @pol.una.py tiene prioridad
    """
    if not correo1:
        return correo2
    if not correo2:
        return correo1
    
    es_institucional_1 = '@pol.una.py' in correo1.lower()
    es_institucional_2 = '@pol.una.py' in correo2.lower()
    
    if es_institucional_1 and not es_institucional_2:
        return correo1
    elif es_institucional_2 and not es_institucional_1:
        return correo2
    else:
        # Si ambos son institucionales o ninguno lo es, mantener el primero
        return correo1

def generar_correo_generico(nombre_completo, correos_existentes=None):
    """
    Genera un correo genérico basado en el nombre completo.
    Formato: primer_nombre.apellido@pol.una.py
    Si ya existe, agrega un número: primer_nombre.apellido2@pol.una.py
    
    Args:
        nombre_completo: Nombre completo del docente (ej: "Juan Pérez")
        correos_existentes: Set o dict con correos ya existentes para evitar duplicados
    
    Returns:
        str: Correo genérico único (ej: "juan.perez@pol.una.py")
    """
    if not nombre_completo:
        return None
    
    if correos_existentes is None:
        correos_existentes = set()
    
    # Normalizar nombre: quitar tildes, convertir a minúsculas
    nombre_normalizado = normalizar_nombre_comparacion(nombre_completo)
    palabras = nombre_normalizado.split()
    
    if len(palabras) < 2:
        # Si no hay al menos nombre y apellido, usar el nombre completo
        base_email = palabras[0] if palabras else "docente"
    else:
        # Usar primer nombre y primer apellido
        primer_nombre = palabras[0]
        primer_apellido = palabras[-1]
        base_email = f"{primer_nombre}.{primer_apellido}"
    
    # Remover caracteres especiales que no son válidos en emails
    base_email = re.sub(r'[^a-z0-9.]', '', base_email)
    
    # Generar correo base
    correo_base = f"{base_email}@pol.una.py"
    
    # Si no existe, retornarlo
    if correo_base not in correos_existentes:
        return correo_base
    
    # Si existe, agregar número hasta encontrar uno disponible
    contador = 2
    while True:
        correo_candidato = f"{base_email}{contador}@pol.una.py"
        if correo_candidato not in correos_existentes:
            return correo_candidato
        contador += 1
        # Limitar a 1000 intentos para evitar loops infinitos
        if contador > 1000:
            import random
            correo_candidato = f"{base_email}{random.randint(1000, 9999)}@pol.una.py"
            return correo_candidato

def es_correo_institucional(correo):
    """
    Verifica si un correo es institucional (@pol.una.py)
    """
    if not correo:
        return False
    return '@pol.una.py' in correo.lower()

def estandarizar_nombre_asignatura(nombre):
    """
    Versión DEFINITIVA:
    1. Limpieza.
    2. Corrección de Ortografía y Mayúsculas (Punto 3).
    3. Romanos.
    4. Formato de Guiones (Punto 3).
    """
    if not nombre: return ""
    
    # 1. Limpieza inicial: Remover texto entre paréntesis y corchetes
    # Elimina (*), (Texto), [Texto], etc.
    nombre = re.sub(r'\s*[\(\[].*?[\)\]]', '', str(nombre))
    # Eliminar espacios extra que pudieran quedar
    nombre = ' '.join(nombre.split())
    
    # 2. DICCIONARIO DE CORRECCIONES (Aquí forzamos la mayúscula)
    correcciones = {
        "sotfware": "Software",  # Typo
        "Sotfware": "Software",  # Typo Capitalizado
        "software": "Software",  # <-- ESTO ARREGLA EL CASO "software" vs "Software"
        "datamining": "Data Mining",
        "Datamining": "Data Mining",
        "Tecnologia": "Tecnología",
        "lenguajes": "Lenguajes",
    }
    
    # Reemplazo palabra por palabra para aplicar correcciones
    palabras = nombre.split()
    palabras_corregidas = [correcciones.get(p, p) for p in palabras]
    nombre = " ".join(palabras_corregidas)

    # Helper para romanos
    def numero_a_romano(numero_str):
        try:
            num = int(numero_str)
            mapa = {
                1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V',
                6: 'VI', 7: 'VII', 8: 'VIII', 9: 'IX', 10: 'X'
            }
            return mapa.get(num, str(num))
        except ValueError:
            return numero_str

    # 3. Lógica de Romanos
    # Caso: Final de línea
    nombre = re.sub(r'\b(\d+)$', lambda m: numero_a_romano(m.group(1)), nombre)
    
    # Caso: Electiva/Optativa en medio
    nombre = re.sub(r'(Electiva|Optativa)\s+(\d+)', 
                    lambda m: f"{m.group(1)} {numero_a_romano(m.group(2))}", 
                    nombre, flags=re.IGNORECASE)

    # 4. ARREGLO DE GUIONES (Punto 3)
    # Separa guiones pegados a romanos: "VII-" se convierte en "VII - "
    nombre = re.sub(r'\b([IVX]+)-', r'\1 - ', nombre)
    
    # Limpieza final por si quedaron dobles espacios
    nombre = ' '.join(nombre.split())

    return nombre

def extraer_primer_nombre_apellido(full_name):
    """
    Extrae el primer nombre y el primer apellido de una cadena.
    Formato esperado: "NOMBRES APELLIDOS" o "APELLIDOS, NOMBRES"
    """
    if not full_name:
        return "Sin Nombre"
    
    # Normalizar espacios y usar Title Case
    full_name = " ".join(full_name.split()).title()
    
    if "," in full_name:
        # Caso: "APELLIDOS, NOMBRES"
        partes_segmento = full_name.split(",", 1)
        apellidos = partes_segmento[0].strip().split()
        nombres = partes_segmento[1].strip().split()
        
        p_nombre = nombres[0] if nombres else ""
        p_apellido = apellidos[0] if apellidos else ""
        return f"{p_nombre} {p_apellido}".strip()
    
    # Caso: "NOMBRES APELLIDOS"
    partes = full_name.split()
    if len(partes) < 2:
        return full_name
        
    # Heurística mejorada:
    # Si hay 4 partes (ej: Juan Carlos Gomez Perez), tomamos 1ra y 3ra.
    if len(partes) == 4:
        return f"{partes[0]} {partes[2]}"
    
    # Si hay 3 partes (ej: Juan Gomez Perez), tomamos 1ra y 2da.
    if len(partes) == 3:
        return f"{partes[0]} {partes[1]}"
        
    # Para cualquier otro caso (2 o 5+), tomamos 1ra y la mitad aproximada o solo 1ra y 2da
    return f"{partes[0]} {partes[1]}"


def es_correo_estudiante(correo):
    """
    Verifica si un correo es de estudiante (@fpuna.edu.py)
    """
    if not correo:
        return False
    return '@fpuna.edu.py' in correo.lower()

def estandarizar_nombre_carrera(carrera):
    """
    Devuelve las carreras con el formato esperado por el programita.
    Formato esperado: Siglas de las carreras separadas por comas (si mas de una).
    """
    # 1. Quita tildes, cambia caracteres blancos a ','
    CARRERAS = ["IIN", "ISP", "IMK", "IEK", "IEL", "IAE", "ICM", "IEN", "LCIK", "LEL", "LCA", "LGH", "LCI"]
    carrera = unicodedata.normalize('NFD', carrera)
    carrera = ''.join(c for c in carrera if unicodedata.category(c) != 'Mn')
    reemplazar = ['-','(',')',' ']
    proceso = carrera.strip()
    for c in reemplazar:
        proceso = proceso.replace(c,',')
    proceso = proceso.split(",")
    out = []

    # 2. Correcion de errores de ortografia comunes, prepara la salida
    for s in proceso:
        if s.upper() in CARRERAS:
            out.append(s.upper())
        elif s.upper()=="LICK":
            out.append("LCIK")
        elif s.upper()=="IIF":
            out.append("IIN")
        elif s.upper()=="LHG":
            out.append("LGH") 
    
    return ",".join(out)

def filtrar_carreras(info, ruta_invalidos="usuarios_no_validos.txt") -> List[List[Any]]:
    """
    Recibe una lista de registros y devuelve la lista filtrada con la carrera normalizada.
    Además gestiona correctamente el archivo de usuarios no válidos.
    Cada registro `r` se asume como: [nombre, correo, ci, carreras]
    """

    def elem_a_str(x):
        if x is None:
            return ""
        return str(x)

    filtrado = []
    invalidos = []

    for r in info:
        # proteger contra registros mal formados
        if len(r) < 4:
            invalidos.append(r)
            continue

        correo = r[1]
        if es_correo_estudiante(correo):
            carrera = estandarizar_nombre_carrera(r[3])
            nombre = " ".join([n.capitalize() for n in r[0].split()]) if r[0] else ""
            if carrera:
                filtrado.append([nombre, correo, r[2], carrera])
            else:
                # correo válido pero no se pudo normalizar la carrera
                invalidos.append(r)
        else:
            # correo no válido
            invalidos.append(r)

    # Escribir todos los usuarios no válidos de una sola vez.
    # Usamos "w" para sobrescribir con el conjunto actual; cambiar a "a" para añadir.
    if invalidos:
        with open(ruta_invalidos, "w", encoding="utf-8", newline="\n") as f:
            for r in invalidos:
                linea = "|".join(elem_a_str(x) for x in r)
                f.write(linea + "\n")

    return filtrado

def matriz_a_txt(matriz: List[List],
    ruta_archivo: str,
    separador: str = "|",
    encabezados: Optional[List[str]] = None,
    encoding: str = "utf-8",
    tratar_none_como_vacio: bool = True
):
    """
    Crea un archivo txt con el contenido de una matriz
    Separador default: "|"
    """
    if not isinstance(matriz, list):
        raise TypeError("El parámetro 'matriz' debe ser una lista de listas.")
    for i, fila in enumerate(matriz):
        if not isinstance(fila, list):
            raise TypeError(f"Cada fila de 'matriz' debe ser una lista. Fila {i} no lo es.")

    def elem_a_str(x):
        if x is None and tratar_none_como_vacio:
            return ""
        return str(x)

    # Abrir archivo y escribir
    with open(ruta_archivo, "w", encoding=encoding, newline="\n") as f:
        # escribir encabezados si se proporcionan
        if encabezados is not None:
            if not isinstance(encabezados, list) or not all(isinstance(h, str) for h in encabezados):
                raise TypeError("Los 'encabezados' deben ser una lista de strings.")
            f.write(separador.join(encabezados) + "\n")

        # escribir cada fila de la matriz
        for fila in matriz:
            # convertir cada elemento a string y unir con el separador
            linea = separador.join(elem_a_str(x) for x in fila)
            f.write(linea + "\n")

def _abrir_dialogo_guardar(title: str = "Guardar archivo de Malla",
                           initialfile: str = "malla.xlsx",
                           defaultextension: str = ".xlsx") -> Optional[str]:
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    ruta = filedialog.asksaveasfilename(
        title=title,
        initialfile=initialfile,
        defaultextension=defaultextension,
        filetypes=[
            ("Archivos de Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    root.destroy()
    return ruta if ruta else None

def guardar_matriz_como_excel(
    matriz: List[List[Any]],
    ruta: Optional[str] = None,
    tiene_encabezado: bool = False,
    sheet_name: str = "Sheet1"
) -> Optional[str]:
    """
    Guarda una matriz (lista de listas) en un archivo Excel usando openpyxl.
    - matriz: lista de filas, cada fila es una lista de valores.
    - ruta: si None, abre diálogo para elegir dónde guardar.
    - tiene_encabezado: si True, la primera fila se usa como encabezado (se escribe igual).
    - Devuelve la ruta donde se guardó o None si se canceló o hubo error.
    """
    # Validaciones básicas
    if not isinstance(matriz, list) or len(matriz) == 0:
        print("❌ La matriz debe ser una lista de listas no vacía.")
        return None

    # Normalizar filas y filtrar None
    filas = [list(f) for f in matriz if f is not None]

    if len(filas) == 0:
        print("❌ No hay filas válidas en la matriz.")
        return None

    # Determinar número máximo de columnas
    max_cols = max(len(r) for r in filas)

    # Rellenar filas cortas con cadenas vacías
    filas_normalizadas = []
    for r in filas:
        if len(r) < max_cols:
            r = list(r) + [""] * (max_cols - len(r))
        filas_normalizadas.append(r)

    # Obtener ruta si no se proporcionó
    if ruta is None:
        ruta = _abrir_dialogo_guardar()
        if ruta is None:
            print("❌ No se seleccionó ninguna ruta de guardado.")
            return None

    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Escribir filas en la hoja
    try:
        for row_idx, fila in enumerate(filas_normalizadas, start=1):
            for col_idx, valor in enumerate(fila, start=1):
                ws.cell(row=row_idx, column=col_idx, value=valor)

        # Asegurar extensión .xlsx
        if not ruta.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            ruta = ruta + ".xlsx"

        wb.save(ruta)
        wb.close()
        print(f"✅ Archivo guardado en: {ruta}")
        return ruta
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {e}")
        return None
