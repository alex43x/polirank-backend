import openpyxl
from tabulate import tabulate
from mysql.connector import Error
import mysql.connector
import os
import tkinter as tk
from tkinter import filedialog
from funciones import insertAsign, insertDoc, insertMalla, insertSecciones
import re

#.\utils\Scripts\venv\Scripts\activate



def procesar_excel_exacto(archivo, nombre_hoja, indices_columnas, fila_inicio):
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        if nombre_hoja not in wb.sheetnames:
            print(f"Error: No existe la hoja '{nombre_hoja}'")
            return
        
        sheet = wb[nombre_hoja]
        data_final = []

        for num_fila, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            # Saltamos hasta llegar a la fila deseada
            if num_fila < fila_inicio:
                continue

            # CORRECCIÓN 1: Inicializamos vacío, sin el num_fila
            fila_procesada = [] 

            # Extraer las columnas específicas por índice
            for i in indices_columnas:
                if i < len(row):
                    fila_procesada.append(row[i])
                else:
                    fila_procesada.append(None)

            # CORRECCIÓN 2: Validamos toda la fila (ya no usamos [1:])
            # Solo agregamos si hay contenido real
            if any(celda is not None for celda in fila_procesada):
                data_final.append(fila_procesada)

        # Mostrar con formato
        if data_final:
            # CORRECCIÓN 3: Quitamos el encabezado de "# Fila"
            '''headers = [f"Col {i}" for i in indices_columnas]
            print(f"\nDatos de '{nombre_hoja}' (Desde fila {fila_inicio}):")
            print(tabulate(data_final, headers=headers, tablefmt="fancy_grid"))'''
            
            return data_final
        else:
            print(f"No hay datos a partir de la fila {fila_inicio}")

    except Exception as e:
        print(f"Error crítico: {e}")

def seleccion_archivo():
    root = tk.Tk()
    root.withdraw() # Ocultamos la ventanita principal

    # TRUCO: Forzar que la ventana aparezca al frente de todo
    root.attributes('-topmost', True) 
    
    print("📂 Abriendo ventana de selección...")
    
    ruta_archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de Malla",
        filetypes=[
            ("Archivos de Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    
    # Importante: Destruimos la instancia de tkinter para liberar memoria
    root.destroy() 

    # Validamos
    if ruta_archivo:
        nombre_archivo = os.path.basename(ruta_archivo)
        print(f"✅ Archivo seleccionado: {nombre_archivo}")
        return ruta_archivo
    else:
        print("❌ No se seleccionó ningún archivo.")
        return None

def obtener_nombre_hoja(ruta_archivo):
    """Muestra las hojas disponibles y permite elegir una."""
    try:
        # Leemos solo la estructura para ser rápidos
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, keep_links=False)
        hojas = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"❌ Error al leer hojas: {e}")
        return None

    if len(hojas) == 1:
        print(f"✅ Hoja única detectada: {hojas[0]}")
        return hojas[0]

    print("\n📋 Hojas disponibles:")
    for i, hoja in enumerate(hojas):
        print(f"  [{i + 1}] {hoja}")

    while True:
        try:
            opcion = int(input("\n👉 Elige el número de la hoja: ")) - 1
            if 0 <= opcion < len(hojas):
                return hojas[opcion]
            print("⚠️ Número inválido.")
        except ValueError:
            print("⚠️ Ingresa un número.")

def obtener_año_periodo(archivo, nombre_hoja):
    direccion = 'D1'
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        

        ws = wb[nombre_hoja]
        año_periodo = ws[direccion].value
        wb.close()

        # Si la celda está vacía, devolvemos valores por defecto
        if not año_periodo:
            return False, False

        texto_limpio = str(año_periodo).upper()
        
        # 1. AÑO (Protegido)
        match = re.search(r'\d{4}', texto_limpio)
        if match:
            year = int(match.group())
        else:
            return False, False
            
        # 2. PERIODO (Con valor por defecto)
        periodo = 1 # Valor predeterminado (asumimos 1 si no dice nada)
        
        if "SEGUNDO" in texto_limpio or "2DO" in texto_limpio:
            periodo = 2
        elif "PRIMER" in texto_limpio or "1ER" in texto_limpio:
            periodo = 1
        else:
            return False, False
            
        return year, periodo

    except Exception as e:
        print(f"❌ Error procesando fechas: {e}")
        return False, False

def limpiar_pantalla():
    # Funciona en Windows (cls) y Linux/Mac (clear)
    os.system('cls' if os.name == 'nt' else 'clear')

def menu_principal():
    while True:
        limpiar_pantalla()
        print("===================================")
        print("   SISTEMA DE GESTIÓN DE MALLA")
        print("===================================")
        print("[1] Insertar Docentes")
        print("[2] Insertar Asignaturas")
        print("[3] Insertar Mallas")
        print("[4] Insertar Secciones y Cursos")
        print("[0] Salir")
        print("===================================")
        
        opcion = input("Selecciona una opción: ")

        if opcion == '1':
            limpiar_pantalla()
            
            print("\nSeleccione archivo")
            ARCHIVO = seleccion_archivo()
            if not ARCHIVO: continue
            HOJA = obtener_nombre_hoja(ARCHIVO)
            
            #Extrae Nombres y Apellidos de los docentes
            FILA_DE_INICIO = 12
            COLUMNAS_OBJETIVO = [12,13]
            
            
            intoData = procesar_excel_exacto(ARCHIVO, HOJA, COLUMNAS_OBJETIVO, FILA_DE_INICIO)
            
            
            insertDoc(connection,intoData)
            
             
            
            
            
            
            
            input("\nPresiona ENTER para continuar...")
            
        elif opcion == '2':
            
            print("\nSeleccione archivo")
            ARCHIVO = seleccion_archivo()
            if not ARCHIVO: continue
            HOJA = obtener_nombre_hoja(ARCHIVO)
            
            #Extrae Asignatura y Departamento
            
            FILA_DE_INICIO = 12
            COLUMNAS_OBJETIVO = [1,2]
            
            
            intoData = procesar_excel_exacto(ARCHIVO, HOJA, COLUMNAS_OBJETIVO, FILA_DE_INICIO)
            
            
            insertAsign(connection,intoData)
            
           
            input("\nPresiona ENTER para continuar...")
            
        elif opcion == '3':
            print("\nSeleccione archivo")
            ARCHIVO = seleccion_archivo()
            if not ARCHIVO: continue
            HOJA = obtener_nombre_hoja(ARCHIVO)
            FILA_DE_INICIO = 0
            COLUMNAS_OBJETIVO = [0,1,2]
            intoData = procesar_excel_exacto(ARCHIVO, HOJA, COLUMNAS_OBJETIVO, FILA_DE_INICIO)
            
            
            #El excel debe tener 3 columnas, la primera de asignatura, la segunda la carrera(siglas) y la tercera el semestre
            insertMalla(connection, intoData)
            
            input("\nPresiona ENTER para continuar...")
            
        elif opcion == '4':
            print("\nSeleccione archivo")
            ARCHIVO = seleccion_archivo()
            if not ARCHIVO: continue 
            HOJA = obtener_nombre_hoja(ARCHIVO)
            
            # --- VALIDACIÓN IMPORTANTE ---
            year, per = obtener_año_periodo(ARCHIVO, HOJA)
            
            if year is False or per is False:
                print("❌ ERROR: No se pudo detectar el Año o Periodo en la celda D1.")
                print("   Verifique que la celda D1 tenga formato '(PRIMER/SEGUNDO) PERIODO (AÑO)'")
                input("Presiona ENTER para volver al menú...")
                continue # Regresa al menú, no intenta insertar
            # -----------------------------

            FILA_DE_INICIO = 12
            # 2=Asignatura, 8=Turno, 9=Seccion, 12=Apellidos, 13=Nombres
            COLUMNAS_OBJETIVO = [2, 8, 9, 12, 13] 
            
            intoData = procesar_excel_exacto(ARCHIVO, HOJA, COLUMNAS_OBJETIVO, FILA_DE_INICIO)

            insertSecciones(connection, intoData, year, per)
            input("\nPresiona ENTER para continuar...")
            
        elif opcion == '5':
            print("\nSeleccione archivo")
            ARCHIVO = seleccion_archivo()
            HOJA = obtener_nombre_hoja(ARCHIVO)
            FILA_DE_INICIO = 0
            COLUMNAS_OBJETIVO = [8]
            #8 = Seccion
            intoData = procesar_excel_exacto(ARCHIVO, HOJA, COLUMNAS_OBJETIVO, FILA_DE_INICIO)
            
            input("\nPresiona ENTER para continuar...")
            
        
        elif opcion == '0':
            print("¡Hasta luego!")
            break
        else:
            input("Opción no válida. Presiona ENTER para intentar de nuevo...")

try:
    connection = mysql.connector.connect(
        host = "localhost",
        port = 3306,
        user = "root",
        password = "Adelant5?",
        db = "polirank_test2"
    )
    if connection.is_connected():
        if __name__ == "__main__":
            menu_principal()
                  
except Error as ex:
    print("Error durante la conexion : {}".format(ex))
finally:
    if connection.is_connected():
        connection.close()
        print("Coneccion Cerrada")